from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Avg, Count, Sum, Q
import io
import json
import re
import urllib.parse
import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Protection, PatternFill, Border, Side
from django.http import HttpResponse, JsonResponse
from django.conf import settings
from django.contrib import messages
from django.views.decorators.http import require_POST

from .models import School, Teacher, Student, Grade, QuarterGrade, QuarterLock, ClassSubject, UserProfile, TeacherProfile
from .forms import LoginForm, SchoolForm, TeacherForm, StudentForm, GradeForm, ClassSubjectForm
from .utils import (
    normalize_class_name, normalize_subject, is_litsey, class_numeric_part,
    default_subjects_for_class, ensure_class_subjects, is_non_graded,
    get_school_number, is_academic_school, official_subjects
)


QUALITATIVE_CHOICES = [
    ('', '---'),
    ('1', 'Ғайб'),
    ('2', 'Ҳозир'),
    ('3', 'Қонеъкунанда'),
    ('4', 'Қаноатбахш'),
    ('5', 'Аъло'),
]


def std_round(value):
    """Round positive numbers using standard half-up rounding."""
    if value is None:
        return None
    return int(value + 0.5)


def calc_quarterly(qmap):
    """Calculate semi-annual, annual, and final grades from a quarter map."""
    calc = {'semi_annual_1': None, 'semi_annual_2': None, 'annual': None, 'final': None}
    if 1 in qmap and qmap[1] is not None and 2 in qmap and qmap[2] is not None:
        calc['semi_annual_1'] = std_round((qmap[1] + qmap[2]) / 2)
    if 3 in qmap and qmap[3] is not None and 4 in qmap and qmap[4] is not None:
        calc['semi_annual_2'] = std_round((qmap[3] + qmap[4]) / 2)
    if calc['semi_annual_1'] is not None and calc['semi_annual_2'] is not None:
        calc['annual'] = std_round((calc['semi_annual_1'] + calc['semi_annual_2']) / 2)
    if calc['annual'] is not None:
        if 'att' in qmap and qmap['att'] is not None:
            calculated_final = std_round((calc['annual'] + qmap['att']) / 2)
            calc['final'] = min(calculated_final, qmap['att'])
        else:
            calc['final'] = calc['annual']
    return calc


def get_date_quarter(date):
    """Map a date to its academic quarter."""
    m = date.month
    if m in (9, 10, 11):
        return 1
    elif m in (12, 1, 2):
        return 2
    elif m in (3, 4, 5):
        return 3
    elif m in (6, 7, 8):
        return 4
    return None


def is_quarter_locked(school, class_name, subject, quarter):
    """Return True if the given quarter is administratively locked."""
    if quarter is None:
        return False
    return QuarterLock.objects.filter(
        school=school,
        class_name=normalize_class_name(class_name),
        subject=normalize_subject(subject),
        quarter=quarter,
        locked=True
    ).exists()


def get_user_school(user):
    if not user or user.is_anonymous:
        return None
    try:
        return user.userprofile.school
    except UserProfile.DoesNotExist:
        try:
            return user.teacherprofile.school
        except TeacherProfile.DoesNotExist:
            return None


def get_user_role(user):
    if not user or user.is_anonymous:
        return settings.ROLE_TEACHER
    try:
        return user.userprofile.role
    except UserProfile.DoesNotExist:
        return settings.ROLE_TEACHER


def has_school_access(user, school):
    if user.is_superuser:
        return True
    role = get_user_role(user)
    if role == settings.ROLE_DIRECTOR:
        return True
    user_school = get_user_school(user)
    return user_school == school


def get_user_profile(user):
    try:
        return user.userprofile
    except UserProfile.DoesNotExist:
        return None


def can_edit_grade_journal(user, school, class_name, subject):
    """Return True if the user may save grades for this school, class and subject."""
    if user.is_superuser:
        return True
    if not has_school_access(user, school):
        return False
    profile = get_user_profile(user)
    if not profile or profile.school != school:
        return False
    role = get_user_role(user)
    # District/department-level director has master access to any school they can reach
    if role == settings.ROLE_DIRECTOR:
        return True
    # School director (principal) can edit anything in their own school
    if role == settings.ROLE_PRINCIPAL:
        return True
    # Zavuchs have full write/edit access to their school's classes
    if role == getattr(settings, 'ROLE_ZAVUCH', 'zavuch') or user.username.startswith('zavuch_'):
        return True
    if role == settings.ROLE_TEACHER:
        # Regular teachers must be explicitly allocated to this ClassSubject in "Тақсимоти дарсҳо"
        return ClassSubject.objects.filter(
            school=school,
            class_name=normalize_class_name(class_name),
            subject=normalize_subject(subject),
            is_active=True,
        ).filter(
            Q(teacher=user) | Q(allocated_teacher__user=user)
        ).exists()
    return False


def can_view_grade_journal(user, school, class_name, subject):
    """Return True if the user may view this grade journal."""
    if user.is_superuser:
        return True
    if not has_school_access(user, school):
        return False
    role = get_user_role(user)
    if role in (settings.ROLE_DIRECTOR, settings.ROLE_PRINCIPAL):
        return True
    return can_edit_grade_journal(user, school, class_name, subject)


def _add_score(totals, key, total, count):
    if total is None or count is None:
        return
    prev = totals.setdefault(key, [0.0, 0])
    prev[0] += float(total)
    prev[1] += int(count)


def calculate_school_rankings():
    """Return all schools ranked by average GPA from daily and quarterly grades (excluding non-graded classes)."""
    totals = {}

    for g in Grade.objects.filter(score__isnull=False).select_related('student'):
        if is_non_graded(g.student.class_name):
            continue
        _add_score(totals, g.student.school_id, g.score, 1)

    for q in QuarterGrade.objects.filter(quarter__in=(1, 2, 3, 4), grade__isnull=False).select_related('student'):
        if is_non_graded(q.class_name):
            continue
        _add_score(totals, q.student.school_id, q.grade, 1)

    for q in QuarterGrade.objects.filter(quarter=0, att_grade__isnull=False).select_related('student'):
        if is_non_graded(q.class_name):
            continue
        _add_score(totals, q.student.school_id, q.att_grade, 1)

    schools = [s for s in School.objects.all() if is_academic_school(s)]
    data = []
    for school in schools:
        total, count = totals.get(school.id, (0.0, 0))
        gpa = round(total / count, 2) if count else 0.0
        data.append({'school': school, 'gpa': gpa})
    data.sort(key=lambda x: x['gpa'], reverse=True)
    rank = 0
    prev_gpa = None
    for item in data:
        if item['gpa'] != prev_gpa:
            rank += 1
            prev_gpa = item['gpa']
        item['rank'] = rank
    return data


def calculate_class_rankings(school_filter=None):
    """Return class rankings with district and school ranks from all grade records."""
    entries = {}
    academic_school_ids = {s.id for s in School.objects.all() if is_academic_school(s)}

    # Seed with every school/class combination that has students (exclude non-graded classes)
    for row in Student.objects.values('school_id', 'school__name', 'class_name').distinct():
        sid = row['school_id']
        if sid not in academic_school_ids:
            continue
        cname = normalize_class_name(row['class_name'])
        if is_non_graded(cname):
            continue
        sname = row['school__name'] or ''
        key = (sid, cname)
        entries[key] = {'school_id': sid, 'school_name': sname, 'class_name': cname, 'total': 0.0, 'count': 0}

    # Daily grades
    for row in Grade.objects.filter(score__isnull=False).values('student__school', 'student__class_name').annotate(total=Sum('score'), count=Count('score')):
        sid = row['student__school']
        if sid not in academic_school_ids:
            continue
        cname = normalize_class_name(row['student__class_name'])
        if is_non_graded(cname):
            continue
        key = (sid, cname)
        if key not in entries:
            entries[key] = {'school_id': row['student__school'], 'school_name': '', 'class_name': cname, 'total': 0.0, 'count': 0}
        entries[key]['total'] += float(row['total'] or 0)
        entries[key]['count'] += int(row['count'])

    # Quarterly grades
    for row in QuarterGrade.objects.filter(quarter__in=(1, 2, 3, 4), grade__isnull=False).values('student__school', 'class_name').annotate(total=Sum('grade'), count=Count('grade')):
        sid = row['student__school']
        if sid not in academic_school_ids:
            continue
        cname = normalize_class_name(row['class_name'])
        if is_non_graded(cname):
            continue
        key = (sid, cname)
        if key not in entries:
            entries[key] = {'school_id': row['student__school'], 'school_name': '', 'class_name': cname, 'total': 0.0, 'count': 0}
        entries[key]['total'] += float(row['total'] or 0)
        entries[key]['count'] += int(row['count'])

    # Attestation grades
    for row in QuarterGrade.objects.filter(quarter=0, att_grade__isnull=False).values('student__school', 'class_name').annotate(total=Sum('att_grade'), count=Count('att_grade')):
        sid = row['student__school']
        if sid not in academic_school_ids:
            continue
        cname = normalize_class_name(row['class_name'])
        if is_non_graded(cname):
            continue
        key = (sid, cname)
        if key not in entries:
            entries[key] = {'school_id': row['student__school'], 'school_name': '', 'class_name': cname, 'total': 0.0, 'count': 0}
        entries[key]['total'] += float(row['total'] or 0)
        entries[key]['count'] += int(row['count'])

    # Resolve any missing school names
    missing_ids = {e['school_id'] for e in entries.values() if not e['school_name']}
    if missing_ids:
        name_map = {s.id: s.name for s in School.objects.filter(id__in=missing_ids)}
        for e in entries.values():
            if not e['school_name']:
                e['school_name'] = name_map.get(e['school_id'], '')

    data = [
        {
            'school_id': e['school_id'],
            'school_name': e['school_name'],
            'class_name': e['class_name'],
            'gpa': round(e['total'] / e['count'], 2) if e['count'] else 0.0,
        }
        for e in entries.values()
    ]
    data.sort(key=lambda x: x['gpa'], reverse=True)

    rank = 0
    prev_gpa = None
    for item in data:
        if item['gpa'] != prev_gpa:
            rank += 1
            prev_gpa = item['gpa']
        item['district_rank'] = rank

    school_state = {}
    for item in data:
        school = item['school_name']
        gpa = item['gpa']
        if school not in school_state:
            school_state[school] = {'rank': 0, 'prev_gpa': None}
        if gpa != school_state[school]['prev_gpa']:
            school_state[school]['rank'] += 1
            school_state[school]['prev_gpa'] = gpa
        item['school_rank'] = school_state[school]['rank']

    if school_filter:
        try:
            sid = int(school_filter)
            data = [item for item in data if item['school_id'] == sid]
        except (ValueError, TypeError):
            data = [item for item in data if school_filter.lower() in item['school_name'].lower()]

    return data


def calculate_subject_rankings():
    """Return subject rankings including all default subjects with 0.0 GPA if no grades exist."""
    totals = {}
    for row in Grade.objects.filter(score__isnull=False).values('subject').annotate(total=Sum('score'), count=Count('score')):
        subj = normalize_subject(row['subject'])
        _add_score(totals, subj, row['total'], row['count'])
    for row in QuarterGrade.objects.filter(quarter__in=(1, 2, 3, 4), grade__isnull=False).values('subject').annotate(total=Sum('grade'), count=Count('grade')):
        subj = normalize_subject(row['subject'])
        _add_score(totals, subj, row['total'], row['count'])
    for row in QuarterGrade.objects.filter(quarter=0, att_grade__isnull=False).values('subject').annotate(total=Sum('att_grade'), count=Count('att_grade')):
        subj = normalize_subject(row['subject'])
        _add_score(totals, subj, row['total'], row['count'])

    # Build a complete set of default subjects from ClassSubject and TJC defaults
    all_subjects = set()
    for subj in ClassSubject.objects.values_list('subject', flat=True).distinct():
        all_subjects.add(normalize_subject(subj))
    for subject_list in settings.TJC_SUBJECTS.values():
        for subj in subject_list:
            all_subjects.add(normalize_subject(subj))

    official = official_subjects()
    data = []
    for subj in all_subjects:
        if subj not in official:
            continue
        total, count = totals.get(subj, (0.0, 0))
        gpa = round(total / count, 2) if count else 0.0
        data.append({'subject': subj, 'gpa': gpa})

    data.sort(key=lambda x: x['gpa'], reverse=True)
    rank = 0
    prev_gpa = None
    for item in data:
        if item['gpa'] != prev_gpa:
            rank += 1
            prev_gpa = item['gpa']
        item['rank'] = rank
    return data


def dashboard(request):
    role = get_user_role(request.user)
    user_school = get_user_school(request.user)

    # Live rankings from Grade and QuarterGrade records
    all_school_ranking = calculate_school_rankings()
    all_subject_ranking = calculate_subject_rankings()

    if not request.user.is_authenticated or request.user.is_superuser or role == settings.ROLE_DIRECTOR:
        schools = School.objects.all()
    else:
        schools = School.objects.filter(id=user_school.id) if user_school else School.objects.none()

    # Full district-wide rankings are visible to all logged-in users
    school_ranking = all_school_ranking
    subject_ranking = all_subject_ranking

    schools_dropdown = [s for s in School.objects.all().order_by('name') if is_academic_school(s)]

    # Optional school filter for class ranking dropdown
    selected_school = request.GET.get('school')
    if selected_school is None:
        # Default to the user's own school on first load
        selected_school = str(user_school.id) if user_school else '0'
    else:
        selected_school = selected_school.strip() or '0'

    class_ranking = calculate_class_rankings()
    if selected_school != '0':
        try:
            sid = int(selected_school)
            class_ranking = [c for c in class_ranking if c['school_id'] == sid]
        except (ValueError, TypeError):
            class_ranking = []

    total_schools = School.objects.count()
    total_students = Student.objects.count()
    total_teachers = Teacher.objects.count()
    try:
        ratio = round(total_students / total_teachers, 1) if total_teachers else 0.0
    except ZeroDivisionError:
        ratio = 0.0

    context = {
        'role': role,
        'schools': schools,
        'schools_dropdown': schools_dropdown,
        'school_ranking': school_ranking,
        'class_ranking': class_ranking,
        'subject_ranking': subject_ranking,
        'total_schools': total_schools,
        'total_students': total_students,
        'total_teachers': total_teachers,
        'ratio': ratio,
        'user_school': user_school,
        'selected_school': selected_school,
    }
    return render(request, 'portal/dashboard.html', context)


def login_view(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user:
                login(request, user)
                return redirect('dashboard')
            else:
                form.add_error(None, 'Номи корбар ёки рамз нодуруст')
    else:
        form = LoginForm()
    return render(request, 'portal/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def school_list(request):
    role = get_user_role(request.user)
    if request.user.is_superuser or role == settings.ROLE_DIRECTOR:
        schools = School.objects.all()
    else:
        user_school = get_user_school(request.user)
        schools = School.objects.filter(id=user_school.id) if user_school else School.objects.none()

    def school_sort_key(school):
        name_lower = school.name.lower()
        type_lower = school.type.lower()

        if 'идор' in type_lower:
            group_priority = 4
        elif 'томактаб' in type_lower:
            group_priority = 3
        elif 'лит' in type_lower or 'лиц' in type_lower:
            group_priority = 2
        elif 'гимн' in name_lower:
            group_priority = 1
            return (group_priority, 0, school.name)
        else:
            group_priority = 1

        nums = re.findall(r'\d+', school.name)
        num = int(nums[0]) if nums else 999999

        return (group_priority, num, school.name)

    schools_list = list(schools)
    schools_list.sort(key=school_sort_key)

    return render(request, 'portal/school_list.html', {'schools': schools_list, 'role': role})


@login_required
def school_add(request):
    if not request.user.is_superuser:
        return redirect('school_list')
    if request.method == 'POST':
        form = SchoolForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('school_list')
    else:
        form = SchoolForm()
    return render(request, 'portal/school_form.html', {'form': form})


def class_list(request, school_id=None):
    user_school = get_user_school(request.user)
    role = get_user_role(request.user)
    if school_id:
        school = get_object_or_404(School, id=school_id)
    else:
        school = user_school
    if request.user.is_authenticated and not has_school_access(request.user, school):
        return redirect('dashboard')

    classes = Student.objects.filter(school=school).values('class_name').distinct().order_by('class_name')
    graded_stats = []
    non_graded_stats = []
    for c in classes:
        cname = c['class_name']
        if is_non_graded(cname):
            label = 'Стикерҳо' if str(class_numeric_part(cname)) == '1' else 'Ғайрибаҳо'
            non_graded_stats.append({'class_name': cname, 'school_rank': '-', 'gpa': label})
            continue

        total = 0.0
        count = 0
        for g in Grade.objects.filter(student__school=school, student__class_name=cname, score__isnull=False):
            total += float(g.score)
            count += 1
        for q in QuarterGrade.objects.filter(student__school=school, class_name=cname, quarter__in=(1, 2, 3, 4), grade__isnull=False):
            total += float(q.grade)
            count += 1
        for q in QuarterGrade.objects.filter(student__school=school, class_name=cname, quarter=0, att_grade__isnull=False):
            total += float(q.att_grade)
            count += 1

        gpa = round(total / count, 2) if count else 0.0
        graded_stats.append({'class_name': cname, 'school_rank': None, 'gpa': gpa})

    # Assign school rank by GPA (descending)
    graded_stats.sort(key=lambda x: x['gpa'], reverse=True)
    for idx, s in enumerate(graded_stats, 1):
        s['school_rank'] = idx

    # Combine in original class_name order
    rank_map = {s['class_name']: s['school_rank'] for s in graded_stats}
    gpa_map = {s['class_name']: s['gpa'] for s in graded_stats}
    class_stats = []
    for c in classes:
        cname = c['class_name']
        if cname in rank_map:
            class_stats.append({'class_name': cname, 'school_rank': rank_map[cname], 'gpa': gpa_map[cname]})
        else:
            for s in non_graded_stats:
                if s['class_name'] == cname:
                    class_stats.append(s)
                    break

    return render(request, 'portal/class_list.html', {'school': school, 'class_stats': class_stats, 'role': role})


def class_detail(request, school_id, class_name):
    school = get_object_or_404(School, id=school_id)
    if request.user.is_authenticated and not has_school_access(request.user, school):
        return redirect('dashboard')
    class_name = normalize_class_name(class_name)
    ensure_class_subjects(school, class_name)
    students = Student.objects.filter(school=school, class_name=class_name).order_by('full_name')
    subjects = ClassSubject.objects.filter(
        school=school, class_name=class_name, is_active=True, subject__in=official_subjects()
    ).order_by('subject')
    non_graded = is_non_graded(class_name)
    return render(request, 'portal/class_detail.html', {
        'school': school,
        'class_name': class_name,
        'students': students,
        'subjects': subjects,
        'non_graded': non_graded,
    })


@login_required
def sticker_entry(request, school_id, class_name):
    """Interactive sticker, attendance and behavior journal for non-graded classes."""
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')

    class_name = normalize_class_name(class_name)
    if not is_non_graded(class_name):
        return redirect('class_detail', school_id=school.id, class_name=class_name)

    subject = normalize_subject('Стикерҳо')
    # Ensure a ClassSubject record exists for permission control
    ClassSubject.objects.get_or_create(
        school=school,
        class_name=class_name,
        subject=subject,
        defaults={'is_default': False, 'is_active': True}
    )

    if not can_view_grade_journal(request.user, school, class_name, subject):
        return HttpResponse('Дастрасӣ манъ аст.', status=403)

    students = Student.objects.filter(school=school, class_name=class_name).order_by('full_name')

    date_str = request.GET.get('date', '')
    try:
        selected_date = datetime.date.fromisoformat(date_str) if date_str else datetime.date.today()
    except ValueError:
        selected_date = datetime.date.today()

    sticker_choices = {'⭐': 'Ситора', '☀️': 'Офтобак', '🌸': 'Гул', '📖': 'Китоб'}

    daily_grades = {}
    daily_attendance = {}
    daily_behavior = {}
    daily_stickers = {}
    for g in Grade.objects.filter(
        student__school=school,
        student__class_name=class_name,
        subject=subject,
        period='Холҳои ҷорӣ (Онлайн)',
        date=selected_date
    ):
        if g.attendance:
            daily_attendance[g.student_id] = g.attendance
        if g.behavior_score is not None:
            daily_behavior[g.student_id] = g.behavior_score
        if g.sticker:
            daily_stickers[g.student_id] = g.sticker

    # Most recent prior behavior score per student (carry-over default)
    prior_behavior = {}
    for g in Grade.objects.filter(
        student__in=students,
        behavior_score__isnull=False,
        date__lt=selected_date
    ).order_by('student_id', '-date'):
        if g.student_id not in prior_behavior:
            prior_behavior[g.student_id] = g.behavior_score

    behavior_default = {}
    for s in students:
        if s.id in daily_behavior:
            behavior_default[s.id] = daily_behavior[s.id]
        elif s.id in prior_behavior:
            behavior_default[s.id] = prior_behavior[s.id]
        else:
            behavior_default[s.id] = 5

    daily_quarter_locked = is_quarter_locked(school, class_name, subject, get_date_quarter(selected_date))

    return render(request, 'portal/sticker_entry.html', {
        'school': school,
        'class_name': class_name,
        'subject': subject,
        'students': students,
        'date': selected_date,
        'daily_attendance': daily_attendance,
        'daily_stickers': daily_stickers,
        'behavior_default': behavior_default,
        'sticker_choices': sticker_choices,
        'daily_quarter_locked': daily_quarter_locked,
    })


@login_required
@require_POST
def add_student(request, school_id, class_name):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')
    class_name = normalize_class_name(class_name)
    full_name = request.POST.get('full_name', '').strip()
    if not full_name:
        messages.error(request, 'Номи хонанда ворид карда шавад.')
        return redirect('class_detail', school_id=school.id, class_name=class_name)
    student_id = f"{school.name}__{class_name}__{full_name}"
    if Student.objects.filter(id=student_id).exists():
        messages.warning(request, 'Хонанда бо ин ном аллакай вуҷуд дорад.')
    else:
        Student.objects.create(
            id=student_id,
            full_name=full_name,
            class_name=class_name,
            school=school
        )
        messages.success(request, 'Хонанда илова шуд.')
    return redirect('class_detail', school_id=school.id, class_name=class_name)


@login_required
@require_POST
def remove_student(request, school_id, class_name):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')
    class_name = normalize_class_name(class_name)
    student_id = request.POST.get('student_id', '').strip()
    if student_id:
        try:
            student = Student.objects.get(id=student_id, school=school, class_name=class_name)
            student.delete()
            messages.success(request, 'Хонанда хориҷ карда шуд.')
        except Student.DoesNotExist:
            messages.error(request, 'Хонанда ёфт нашуд.')
    return redirect('class_detail', school_id=school.id, class_name=class_name)


@login_required
@require_POST
def edit_student(request, school_id, class_name):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')
    class_name = normalize_class_name(class_name)
    student_id = request.POST.get('student_id', '').strip()
    full_name = request.POST.get('full_name', '').strip()

    if not student_id or not full_name:
        messages.error(request, 'Иттилооти нокифоя барои таҳрири хонанда.')
        return redirect('class_detail', school_id=school.id, class_name=class_name)

    try:
        student = Student.objects.get(id=student_id, school=school, class_name=class_name)
    except Student.DoesNotExist:
        messages.error(request, 'Хонанда ёфт нашуд.')
        return redirect('class_detail', school_id=school.id, class_name=class_name)

    student.full_name = full_name
    student.save()
    messages.success(request, f'Хонанда {full_name} таҳрир шуд.')
    return redirect('class_detail', school_id=school.id, class_name=class_name)


@login_required
def grade_entry(request, school_id, class_name, subject):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')

    class_name = normalize_class_name(class_name)
    subject = normalize_subject(subject)

    if not can_view_grade_journal(request.user, school, class_name, subject):
        return HttpResponse('Дастрасӣ манъ аст.', status=403)

    students = Student.objects.filter(school=school, class_name=class_name).order_by('full_name')

    if request.method == 'POST':
        if not can_edit_grade_journal(request.user, school, class_name, subject):
            return HttpResponse('Дастрасӣ барои тағйир додан манъ аст.', status=403)

        date_str = request.POST.get('date', '')
        try:
            date = datetime.date.fromisoformat(date_str) if date_str else datetime.date.today()
        except ValueError:
            date = datetime.date.today()

        daily_q = get_date_quarter(date)
        daily_locked = is_quarter_locked(school, class_name, subject, daily_q)

        for student in students:
            # daily grade fallback (skip if quarter is locked)
            if not daily_locked:
                grade, _ = Grade.objects.get_or_create(
                    student=student,
                    subject=subject,
                    period='Холҳои ҷорӣ (Онлайн)',
                    date=date,
                    defaults={'score': None, 'attendance': None, 'behavior_score': None}
                )

                key = f"score_{student.id}"
                val = request.POST.get(key, '').strip()
                if val:
                    try:
                        score = float(val)
                        if not score.is_integer():
                            score = None
                        else:
                            score = int(score)
                            if not (1 <= score <= 10):
                                score = None
                    except ValueError:
                        score = None
                else:
                    score = None
                grade.score = score

                att = request.POST.get(f"attendance_{student.id}", '').strip()
                grade.attendance = att if att in ('+', '-') else None

                beh = request.POST.get(f"behavior_{student.id}", '').strip()
                if beh:
                    try:
                        b = int(beh)
                        grade.behavior_score = b if 1 <= b <= 5 else None
                    except ValueError:
                        grade.behavior_score = None
                else:
                    grade.behavior_score = None

                if grade.score is None and not grade.attendance and grade.behavior_score is None:
                    grade.delete()
                else:
                    grade.save()

            # quarterly grade fallback (skip locked quarters)
            for q in (1, 2, 3, 4, 0):
                if is_quarter_locked(school, class_name, subject, q):
                    continue
                if q == 0:
                    qkey = f"att_{student.id}"
                else:
                    qkey = f"q_{q}_{student.id}"
                qval = request.POST.get(qkey, '').strip()
                if qval:
                    try:
                        qscore = float(qval)
                        if 1 <= qscore <= 10:
                            if q == 0:
                                QuarterGrade.objects.update_or_create(
                                    student=student,
                                    class_name=class_name,
                                    subject=subject,
                                    quarter=0,
                                    defaults={'att_grade': int(qscore)}
                                )
                            else:
                                QuarterGrade.objects.update_or_create(
                                    student=student,
                                    class_name=class_name,
                                    subject=subject,
                                    quarter=q,
                                    defaults={'grade': int(qscore)}
                                )
                    except ValueError:
                        pass
                else:
                    if q == 0:
                        QuarterGrade.objects.filter(
                            student=student, class_name=class_name, subject=subject, quarter=0
                        ).delete()
                    else:
                        QuarterGrade.objects.filter(
                            student=student, class_name=class_name, subject=subject, quarter=q
                        ).delete()
        return redirect('grade_entry', school_id=school.id, class_name=class_name, subject=subject)

    date_str = request.GET.get('date', '')
    try:
        selected_date = datetime.date.fromisoformat(date_str) if date_str else datetime.date.today()
    except ValueError:
        selected_date = datetime.date.today()

    daily_grades = {}
    daily_attendance = {}
    daily_behavior = {}
    for g in Grade.objects.filter(
        student__school=school,
        student__class_name=class_name,
        subject=subject,
        period='Холҳои ҷорӣ (Онлайн)',
        date=selected_date
    ):
        if g.score is not None:
            daily_grades[g.student_id] = int(g.score) if g.score.is_integer() else g.score
        if g.attendance:
            daily_attendance[g.student_id] = g.attendance
        if g.behavior_score is not None:
            daily_behavior[g.student_id] = g.behavior_score

    # Most recent prior behavior score per student (carry-over default)
    prior_behavior = {}
    for g in Grade.objects.filter(
        student__in=students,
        behavior_score__isnull=False,
        date__lt=selected_date
    ).order_by('student_id', '-date'):
        if g.student_id not in prior_behavior:
            prior_behavior[g.student_id] = g.behavior_score

    behavior_default = {}
    for s in students:
        if s.id in daily_behavior:
            behavior_default[s.id] = daily_behavior[s.id]
        elif s.id in prior_behavior:
            behavior_default[s.id] = prior_behavior[s.id]
        else:
            behavior_default[s.id] = 5

    q_grades = QuarterGrade.objects.filter(
        student__school=school,
        class_name=class_name,
        subject=subject
    ).select_related('student')
    qmap_by_student = {}
    for g in q_grades:
        sid = g.student_id
        if sid not in qmap_by_student:
            qmap_by_student[sid] = {}
        if g.quarter == 0 and g.att_grade is not None:
            qmap_by_student[sid]['att'] = g.att_grade
        elif g.grade is not None:
            qmap_by_student[sid][g.quarter] = g.grade

    quarter_grades = {}
    calculated = {}
    for student in students:
        qmap = qmap_by_student.get(student.id, {})
        quarter_grades[student.id] = qmap
        calculated[student.id] = calc_quarterly(qmap)

    non_graded = is_non_graded(class_name)

    locked_quarters = set(
        QuarterLock.objects.filter(
            school=school,
            class_name=class_name,
            subject=subject,
            locked=True
        ).values_list('quarter', flat=True)
    )
    daily_quarter_locked = is_quarter_locked(school, class_name, subject, get_date_quarter(selected_date))

    return render(request, 'portal/grade_entry.html', {
        'school': school,
        'class_name': class_name,
        'subject': subject,
        'students': students,
        'date': selected_date,
        'daily_grades': daily_grades,
        'daily_attendance': daily_attendance,
        'daily_behavior': daily_behavior,
        'behavior_default': behavior_default,
        'quarter_grades': quarter_grades,
        'calculated': calculated,
        'non_graded': non_graded,
        'qualitative_choices': QUALITATIVE_CHOICES,
        'quarter_numbers': [1, 2, 3, 4],
        'locked_quarters': locked_quarters,
        'daily_quarter_locked': daily_quarter_locked,
    })


@login_required
def add_remove_subject(request, school_id, class_name):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')
    class_name = normalize_class_name(class_name)

    if request.method == 'POST':
        action = request.POST.get('action')
        subject = request.POST.get('subject', '').strip()
        subject = normalize_subject(subject)
        if action == 'add' and subject:
            ClassSubject.objects.get_or_create(
                school=school,
                class_name=class_name,
                subject=subject,
                defaults={'is_active': True, 'is_default': False}
            )
        elif action == 'remove' and subject:
            ClassSubject.objects.filter(
                school=school,
                class_name=class_name,
                subject=subject
            ).update(is_active=False)
    return redirect('class_detail', school_id=school.id, class_name=class_name)


@login_required
def download_template(request, class_name=None, school_id=None):
    if class_name:
        class_name = normalize_class_name(class_name)

    if school_id:
        school = get_object_or_404(School, id=school_id)
    else:
        school = get_user_school(request.user)
    school_number = get_school_number(school) if school else 'unknown'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Хонандагон'

    headers = ['№ мактаб', '№ синф', 'Ному насаб', 'Синф']
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for i in range(2, 1001):
        ws.cell(row=i, column=1, value=f'=IF(C{i}<>"", COUNTA($C$2:C{i}), "")')
        ws.cell(row=i, column=2, value=f'=IF(C{i}<>"", COUNTIF($D$2:D{i}, D{i}), "")')

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 15

    # Unlock C (Ному насаб) and D (Синф) for data entry; A and B stay locked
    for row in ws.iter_rows(min_row=2, max_row=1000, min_col=1, max_col=4):
        for cell in row:
            if cell.column in (3, 4):
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.protection = Protection(locked=False)
                cell.number_format = '@'
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.protection.sheet = True
    ws.protection.set_password('maorif_zafarobod')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Шаблони_Синфҳо_Мактаби_{school_number}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    response['Content-Disposition'] = f"attachment; filename*=utf-8''{encoded_filename}"
    return response


@login_required
def import_excel(request, class_name=None, school_id=None):
    if school_id is not None:
        school = get_object_or_404(School, id=school_id)
    else:
        user_school = get_user_school(request.user)
        if request.user.is_superuser and not user_school:
            class_name = normalize_class_name(class_name or '')
            sample_student = Student.objects.filter(class_name=class_name).first()
            if sample_student:
                school = sample_student.school
            else:
                school = School.objects.filter(id=24).first() or School.objects.first()
        else:
            school = user_school

    if not school:
        return redirect('dashboard')

    if request.method != 'POST' or 'excel' not in request.FILES:
        return redirect('class_list', school_id=school.id)

    target_class = normalize_class_name(class_name) if class_name else None
    df = pd.read_excel(request.FILES['excel'])
    df.columns = [str(c).strip() for c in df.columns]

    name_col = 'Ному насаб' if 'Ному насаб' in df.columns else None
    class_col = 'Синф' if 'Синф' in df.columns else None
    if name_col is None:
        messages.error(request, 'Сутуни "Ному насаб" ёфт нашуд.')
        return redirect('class_list', school_id=school.id)

    fixed_cols = {name_col, class_col} if class_col else {name_col}

    imported = 0
    for _, row in df.iterrows():
        full_name = str(row.get(name_col, '')).strip()
        if not full_name or full_name.lower() in ('nan', 'none'):
            continue

        if class_col:
            c_name = str(row.get(class_col, '')).strip()
            if not c_name or c_name.lower() in ('nan', 'none'):
                c_name = target_class
            if not c_name:
                continue
            c_name = normalize_class_name(c_name)
        else:
            if not target_class:
                continue
            c_name = target_class

        student_id = f"{school.name}__{c_name}__{full_name}"
        student, _ = Student.objects.update_or_create(
            id=student_id,
            defaults={'full_name': full_name, 'class_name': c_name, 'school': school}
        )

        for col in df.columns:
            if col in fixed_cols:
                continue
            low = col.lower()
            if 'unnamed' in low or 'жами' in low or 'рейтинг' in low or '№' in col or col.strip() == '':
                continue
            subj = normalize_subject(col)
            val = row.get(col)
            if val is None or str(val).lower() in ('nan', 'none', ''):
                continue
            try:
                score = float(val)
                if 1 <= score <= 10:
                    Grade.objects.update_or_create(
                        student=student,
                        subject=subj,
                        period='Холҳои ҷорӣ (Онлайн)',
                        defaults={'score': score}
                    )
                    imported += 1
            except (ValueError, TypeError):
                continue

    messages.success(request, f'{imported} хол(ҳо) ворид карда шуд.')
    if target_class:
        return redirect('class_detail', school_id=school.id, class_name=target_class)
    return redirect('class_list', school_id=school.id)


@login_required
@require_POST
def save_grade_ajax(request):
    try:
        student_id = request.POST.get('student_id', '').strip()
        subject = normalize_subject(request.POST.get('subject', ''))
        date_str = request.POST.get('date', '').strip()
        score_raw = request.POST.get('score', '')
        if score_raw is not None:
            score_raw = score_raw.strip().replace(',', '.')
            if score_raw == '':
                score_raw = None
        grade_type = request.POST.get('type', 'daily')
        quarter_str = request.POST.get('quarter', '').strip()

        if not student_id or not subject:
            return JsonResponse({'success': False, 'message': 'Далелҳои нокифоя.'}, status=200)

        try:
            student = Student.objects.get(id=student_id)
        except Student.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Хонанда ёфт нашуд.'}, status=200)

        class_name = normalize_class_name(student.class_name)

        if not can_edit_grade_journal(request.user, student.school, class_name, subject):
            return JsonResponse({'success': False, 'message': 'Дастрасӣ барои тағйир додан манъ аст.'}, status=403)

        def parse_score(raw):
            if not raw:
                return None
            try:
                score = float(raw)
            except ValueError:
                raise ValueError('Хол бояд рақам бошад.')
            return int(score)

        if grade_type == 'daily':
            try:
                date = datetime.date.fromisoformat(date_str) if date_str else datetime.date.today()
            except ValueError:
                return JsonResponse({'success': False, 'message': 'Санаи нодуруст.'}, status=200)

            if is_quarter_locked(student.school, class_name, subject, get_date_quarter(date)):
                return JsonResponse({'success': False, 'message': 'Ин чоряк баста шудааст.'}, status=200)

            grade, _ = Grade.objects.get_or_create(
                student=student,
                subject=subject,
                period='Холҳои ҷорӣ (Онлайн)',
                date=date,
                defaults={'score': None, 'attendance': None, 'behavior_score': None, 'sticker': None}
            )

            if 'score' in request.POST:
                raw = request.POST.get('score', '').strip().replace(',', '.')
                if raw == '':
                    raw = None
                try:
                    score = parse_score(raw)
                except ValueError as e:
                    return JsonResponse({'success': False, 'message': str(e)}, status=200)
                grade.score = score

            if 'attendance' in request.POST:
                att = request.POST.get('attendance', '').strip()
                grade.attendance = att if att in ('+', '-') else None

            if 'behavior_score' in request.POST:
                beh_raw = request.POST.get('behavior_score', '').strip()
                if beh_raw:
                    try:
                        b = int(beh_raw)
                        grade.behavior_score = b if 1 <= b <= 5 else None
                    except ValueError:
                        grade.behavior_score = None
                else:
                    grade.behavior_score = None

            if 'sticker' in request.POST:
                st = request.POST.get('sticker', '').strip()
                grade.sticker = st if st in ('⭐', '☀️', '🌸', '📖') else None

            if grade.score is None and not grade.attendance and grade.behavior_score is None and not grade.sticker:
                grade.delete()
            else:
                grade.save()
            return JsonResponse({'success': True, 'saved': True}, status=200)

        elif grade_type == 'quarterly':
            try:
                quarter = int(quarter_str)
            except ValueError:
                return JsonResponse({'success': False, 'message': 'Чораки нодуруст.'}, status=200)

            if is_quarter_locked(student.school, class_name, subject, quarter):
                return JsonResponse({'success': False, 'message': 'Ин чоряк баста шудааст.'}, status=200)

            try:
                score = parse_score(score_raw)
            except ValueError as e:
                return JsonResponse({'success': False, 'message': str(e)}, status=200)

            if score is not None and 1 <= score <= 10:
                if quarter == 0:
                    QuarterGrade.objects.update_or_create(
                        student=student,
                        class_name=class_name,
                        subject=subject,
                        quarter=0,
                        defaults={'att_grade': score}
                    )
                else:
                    QuarterGrade.objects.update_or_create(
                        student=student,
                        class_name=class_name,
                        subject=subject,
                        quarter=quarter,
                        defaults={'grade': score}
                    )
            else:
                QuarterGrade.objects.filter(
                    student=student,
                    class_name=class_name,
                    subject=subject,
                    quarter=quarter
                ).delete()

            qmap = {}
            for g in QuarterGrade.objects.filter(student=student, class_name=class_name, subject=subject):
                if g.quarter == 0 and g.att_grade is not None:
                    qmap['att'] = g.att_grade
                elif g.grade is not None:
                    qmap[g.quarter] = g.grade
            calc = calc_quarterly(qmap)
            return JsonResponse({'success': True, 'saved': True, **calc}, status=200)

        return JsonResponse({'success': False, 'message': 'Навъи нодуруст.'}, status=200)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=200)


@login_required
@require_POST
def calc_quarter_from_daily(request, school_id, class_name, subject):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return HttpResponse('Дастрасӣ манъ аст.', status=403)
    class_name = normalize_class_name(class_name)
    subject = normalize_subject(subject)

    if not can_edit_grade_journal(request.user, school, class_name, subject):
        return HttpResponse('Дастрасӣ барои тағйир додан манъ аст.', status=403)

    students = Student.objects.filter(school=school, class_name=class_name)
    for student in students:
        grades = Grade.objects.filter(
            student=student,
            subject=subject,
            period='Холҳои ҷорӣ (Онлайн)',
            score__isnull=False
        )
        sums = {1: [0.0, 0], 2: [0.0, 0], 3: [0.0, 0], 4: [0.0, 0]}
        for g in grades:
            m = g.date.month
            if m in (9, 10, 11):
                q = 1
            elif m in (12, 1, 2):
                q = 2
            elif m in (3, 4, 5):
                q = 3
            elif m in (6, 7, 8):
                q = 4
            else:
                continue
            sums[q][0] += g.score
            sums[q][1] += 1
        for q, (s, c) in sums.items():
            if c and not is_quarter_locked(school, class_name, subject, q):
                avg = std_round(s / c)
                QuarterGrade.objects.update_or_create(
                    student=student,
                    class_name=class_name,
                    subject=subject,
                    quarter=q,
                    defaults={'grade': avg}
                )
    return redirect('grade_entry', school_id=school.id, class_name=class_name, subject=subject)


def _student_subject_scores(student):
    """Return a dict of {normalized_subject: average_score} for a student."""
    totals = {}
    for row in Grade.objects.filter(student=student, score__isnull=False).values('subject').annotate(total=Sum('score'), count=Count('score')):
        subj = normalize_subject(row['subject'])
        _add_score(totals, subj, row['total'], row['count'])
    for row in QuarterGrade.objects.filter(student=student, quarter__in=(1, 2, 3, 4), grade__isnull=False).values('subject').annotate(total=Sum('grade'), count=Count('grade')):
        subj = normalize_subject(row['subject'])
        _add_score(totals, subj, row['total'], row['count'])
    for row in QuarterGrade.objects.filter(student=student, quarter=0, att_grade__isnull=False).values('subject').annotate(total=Sum('att_grade'), count=Count('att_grade')):
        subj = normalize_subject(row['subject'])
        _add_score(totals, subj, row['total'], row['count'])
    return {subj: round(total / count, 2) if count else 0.0 for subj, (total, count) in totals.items()}


def _all_student_gpas():
    """Return {student_id: overall_gpa} for every student using Grade and QuarterGrade."""
    totals = {}
    for row in Grade.objects.filter(score__isnull=False).values('student_id').annotate(total=Sum('score'), count=Count('score')):
        _add_score(totals, row['student_id'], row['total'], row['count'])
    for row in QuarterGrade.objects.filter(quarter__in=(1, 2, 3, 4), grade__isnull=False).values('student_id').annotate(total=Sum('grade'), count=Count('grade')):
        _add_score(totals, row['student_id'], row['total'], row['count'])
    for row in QuarterGrade.objects.filter(quarter=0, att_grade__isnull=False).values('student_id').annotate(total=Sum('att_grade'), count=Count('att_grade')):
        _add_score(totals, row['student_id'], row['total'], row['count'])

    gpas = {}
    for student in Student.objects.all():
        total, count = totals.get(student.id, (0.0, 0))
        gpas[student.id] = round(total / count, 2) if count else 0.0
    return gpas


def student_detail(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if request.user.is_authenticated and not has_school_access(request.user, student.school):
        return redirect('dashboard')

    subject_scores = _student_subject_scores(student)
    subjects = sorted(subject_scores.keys())
    scores = [subject_scores[s] for s in subjects]

    all_gpas = _all_student_gpas()
    student_gpa = all_gpas.get(student.id, 0.0)

    class_gpas = []
    school_gpas = []
    district_gpas = []
    for s in Student.objects.all():
        g = all_gpas.get(s.id, 0.0)
        district_gpas.append(g)
        if s.school_id == student.school_id:
            school_gpas.append(g)
        if s.school_id == student.school_id and s.class_name == student.class_name:
            class_gpas.append(g)

    class_rank = len({g for g in class_gpas if g > student_gpa}) + 1
    school_rank = len({g for g in school_gpas if g > student_gpa}) + 1
    district_rank = len({g for g in district_gpas if g > student_gpa}) + 1

    non_graded = is_non_graded(student.class_name)

    recent_stickers = Grade.objects.filter(
        student=student,
        sticker__isnull=False
    ).order_by('-date')[:15]

    context = {
        'student': student,
        'gpa': student_gpa,
        'non_graded': non_graded,
        'class_rank': class_rank,
        'school_rank': school_rank,
        'district_rank': district_rank,
        'subjects': subjects,
        'scores': scores,
        'recent_stickers': recent_stickers,
        'sticker_labels': {'⭐': 'Ситора', '☀️': 'Офтобак', '🌸': 'Гул', '📖': 'Китоб'},
    }
    return render(request, 'portal/student_detail.html', context)


def _next_teacher_counter(school_num):
    """Return the next sequential teacher username counter for a school."""
    prefix = f'teacher_{school_num}_'
    pattern = re.compile(rf'^{re.escape(prefix)}(\d+)$')
    max_counter = 0
    for username in User.objects.filter(username__startswith=prefix).values_list('username', flat=True):
        m = pattern.match(username)
        if m:
            max_counter = max(max_counter, int(m.group(1)))
    return max_counter + 1


@login_required
def teacher_list(request, school_id=None):
    role = get_user_role(request.user)
    user_school = get_user_school(request.user)
    if school_id:
        school = get_object_or_404(School, id=school_id)
    else:
        school = user_school
    if not has_school_access(request.user, school):
        return redirect('dashboard')
    teachers = Teacher.objects.filter(school=school).order_by('name')
    for teacher in teachers:
        tp = TeacherProfile.objects.filter(school=school, full_name=teacher.name).select_related('user').first()
        if tp and tp.user:
            teacher.username = tp.user.username
            teacher.is_password_private = tp.user.last_login is not None
            if teacher.is_password_private:
                teacher.password_display = 'Рамзи шахсӣ 🔒'
            else:
                teacher.password_display = f'{tp.user.username.capitalize()}@2026'
        else:
            teacher.username = ''
            teacher.is_password_private = False
            teacher.password_display = '—'
    return render(request, 'portal/teacher_list.html', {'school': school, 'teachers': teachers, 'role': role})


@login_required
@require_POST
def add_teacher(request, school_id):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')

    full_name = request.POST.get('full_name', '').strip()
    phone = request.POST.get('phone', '').strip()
    subject = request.POST.get('subject', '').strip()

    if not full_name:
        messages.error(request, 'Ному насаби омӯзгор бояд ворид карда шавад.')
        return redirect('teacher_list', school_id=school.id)

    school_num = get_school_number(school)
    counter = _next_teacher_counter(school_num)
    username = f'teacher_{school_num}_{counter}'
    password = f'Teacher_{school_num}_{counter}@2026'

    try:
        user = User.objects.create_user(
            username=username,
            password=password,
            first_name=full_name,
        )
    except Exception:
        messages.error(request, 'Эҷоди ҳисоби корбар муваффақият амалӣ нагашт.')
        return redirect('teacher_list', school_id=school.id)

    TeacherProfile.objects.create(
        user=user,
        school=school,
        full_name=full_name,
        phone=phone,
        specialty=subject,
    )
    UserProfile.objects.update_or_create(
        user=user,
        defaults={
            'role': settings.ROLE_TEACHER,
            'school': school,
        },
    )
    Teacher.objects.create(
        school=school,
        name=full_name,
        subject=subject,
        phone=phone,
    )

    messages.success(request, f'Омӯзгор {full_name} бо муваффақият илова шуд. Логин: {username}')
    return redirect('teacher_list', school_id=school.id)


@login_required
@require_POST
def remove_teacher(request, school_id):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')

    teacher_id = request.POST.get('teacher_id', '').strip()
    if not teacher_id:
        messages.error(request, 'ID-и омӯзгор муайян карда нашуд.')
        return redirect('teacher_list', school_id=school.id)

    try:
        teacher = Teacher.objects.get(id=teacher_id, school=school)
    except Teacher.DoesNotExist:
        messages.error(request, 'Омӯзгор ёфт нашуд.')
        return redirect('teacher_list', school_id=school.id)

    full_name = teacher.name
    teacher_profile = TeacherProfile.objects.filter(school=school, full_name=full_name).first()
    if teacher_profile:
        teacher_profile.user.delete()
    teacher.delete()

    messages.success(request, f'Омӯзгор {full_name} хориҷ карда шуд.')
    return redirect('teacher_list', school_id=school.id)


@login_required
@require_POST
def edit_teacher(request, school_id):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        return redirect('dashboard')

    teacher_id = request.POST.get('teacher_id', '').strip()
    full_name = request.POST.get('full_name', '').strip()
    phone = request.POST.get('phone', '').strip()
    subject = request.POST.get('subject', '').strip()
    new_password = request.POST.get('new_password', '').strip()

    if not teacher_id or not full_name:
        messages.error(request, 'Иттилооти нокифоя барои таҳрири омӯзгор.')
        return redirect('teacher_list', school_id=school.id)

    try:
        teacher = Teacher.objects.get(id=teacher_id, school=school)
    except Teacher.DoesNotExist:
        messages.error(request, 'Омӯзгор ёфт нашуд.')
        return redirect('teacher_list', school_id=school.id)

    old_name = teacher.name
    teacher.name = full_name
    teacher.phone = phone
    teacher.subject = subject
    teacher.save()

    teacher_profile = TeacherProfile.objects.filter(school=school, full_name=old_name).first()
    if teacher_profile:
        teacher_profile.full_name = full_name
        teacher_profile.phone = phone
        teacher_profile.specialty = subject
        teacher_profile.save()
        if teacher_profile.user:
            teacher_profile.user.first_name = full_name
            if new_password:
                teacher_profile.user.set_password(new_password)
            teacher_profile.user.save()

    messages.success(request, f'Омӯзгор {full_name} таҳрир шуд.')
    return redirect('teacher_list', school_id=school.id)


@login_required
def download_teacher_template(request, school_id):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        messages.error(request, 'Дастрасӣ ба ин муассиса манъ аст.')
        return redirect('dashboard')
    school_number = get_school_number(school)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Омӯзгорон'

    headers = ['№', 'Ному насаби омӯзгор', 'Рақами телефон', 'Маълумот', 'Ихтисос']
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Sample data row (B-E unlocked for editing; A is formula-driven and locked)
    ws.cell(row=2, column=2, value='Намуна: Раҷабов А.')
    ws.cell(row=2, column=3, value='+992901234567')
    ws.cell(row=2, column=4, value='Олии педагогӣ')
    ws.cell(row=2, column=5, value='Математика')

    for i in range(2, 1001):
        ws.cell(row=i, column=1, value=f'=IF(B{i}<>"", COUNTA($B$2:B{i}), "")')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25

    # Leave A locked with formulas; unlock B, C, D, E for data entry
    for row in ws.iter_rows(min_row=2, max_row=1000, min_col=2, max_col=5):
        for cell in row:
            cell.protection = Protection(locked=False)
            cell.number_format = '@'

    ws.protection.sheet = True
    ws.protection.set_password('maorif_zafarobod')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Шаблони_Омӯзгорон_Мактаби_{school_number}.xlsx"
    encoded_filename = urllib.parse.quote(filename)
    response['Content-Disposition'] = f"attachment; filename*=utf-8''{encoded_filename}"
    return response


@login_required
def import_teachers(request, school_id):
    school = get_object_or_404(School, id=school_id)
    if not has_school_access(request.user, school):
        messages.error(request, 'Дастрасӣ ба ин муассиса манъ аст.')
        return redirect('dashboard')

    if request.method != 'POST' or 'excel' not in request.FILES:
        return redirect('teacher_list', school_id=school.id)

    wb = load_workbook(request.FILES['excel'], data_only=True)
    ws = wb.active

    created_teachers = []
    updated_teachers = []
    school_num = get_school_number(school)
    counter = _next_teacher_counter(school_num)

    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = str(row[1] or '').strip() if len(row) > 1 else ''
        if not full_name or full_name.lower() in ('nan', 'none', ''):
            continue

        phone = str(row[2] or '').strip() if len(row) > 2 else ''
        education = str(row[3] or '').strip() if len(row) > 3 else ''
        specialty = str(row[4] or '').strip() if len(row) > 4 else ''

        # Avoid creating duplicates: update existing teacher details if found.
        existing_tp = TeacherProfile.objects.filter(school=school, full_name=full_name).first()
        existing_teacher = Teacher.objects.filter(school=school, name=full_name).first()
        if existing_tp or existing_teacher:
            if existing_tp:
                existing_tp.phone = phone
                existing_tp.education = education
                existing_tp.specialty = specialty
                existing_tp.save()
                if existing_tp.user:
                    existing_tp.user.first_name = full_name
                    existing_tp.user.save()
            if existing_teacher:
                existing_teacher.phone = phone
                existing_teacher.subject = specialty
                existing_teacher.save()
            updated_teachers.append({'full_name': full_name})
            continue

        # Find the next available username for this school.
        username = f'teacher_{school_num}_{counter}'
        password = f'Teacher_{school_num}_{counter}@2026'
        while User.objects.filter(username=username).exists():
            counter += 1
            username = f'teacher_{school_num}_{counter}'
            password = f'Teacher_{school_num}_{counter}@2026'

        try:
            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=full_name,
            )
        except Exception:
            continue

        TeacherProfile.objects.create(
            user=user,
            school=school,
            full_name=full_name,
            phone=phone,
            education=education,
            specialty=specialty,
        )

        UserProfile.objects.update_or_create(
            user=user,
            defaults={
                'role': settings.ROLE_TEACHER,
                'school': school,
            },
        )

        Teacher.objects.create(
            school=school,
            name=full_name,
            subject=specialty,
            phone=phone,
        )

        created_teachers.append({
            'full_name': full_name,
            'username': username,
            'password': password,
        })
        counter += 1

    return render(request, 'portal/imported_teachers_list.html', {
        'school': school,
        'created_teachers': created_teachers,
        'updated_teachers': updated_teachers,
    })


@login_required
def lesson_allocation(request):
    school = get_user_school(request.user)
    if not school:
        messages.error(request, 'Муассисаи шумо муайян карда нашуд.')
        return redirect('dashboard')
    if not has_school_access(request.user, school):
        return redirect('dashboard')

    class_subjects = ClassSubject.objects.filter(
        school=school, is_active=True
    ).select_related('allocated_teacher', 'teacher').order_by('class_name', 'subject')
    teachers = TeacherProfile.objects.filter(school=school).select_related('user').order_by('full_name')

    user_to_profile = {tp.user_id: tp.id for tp in teachers}
    for cs in class_subjects:
        cs.selected_profile_id = cs.allocated_teacher_id or user_to_profile.get(cs.teacher_id)

    return render(request, 'school/lesson_allocation.html', {
        'school': school,
        'class_subjects': class_subjects,
        'teachers': teachers,
    })


@login_required
@require_POST
def save_lesson_allocation(request):
    school = get_user_school(request.user)
    if not school or not has_school_access(request.user, school):
        return redirect('dashboard')

    cs_keys = [k for k in request.POST if k.startswith('alloc_')]
    cs_ids = []
    for k in cs_keys:
        try:
            cs_ids.append(int(k.split('_', 1)[1]))
        except (ValueError, IndexError):
            continue

    class_subject_map = {
        cs.id: cs for cs in ClassSubject.objects.filter(id__in=cs_ids, school=school)
    }
    teacher_map = {
        tp.id: tp for tp in TeacherProfile.objects.filter(school=school)
    }

    updated = []
    for key in cs_keys:
        try:
            cs_id = int(key.split('_', 1)[1])
        except (ValueError, IndexError):
            continue
        cs = class_subject_map.get(cs_id)
        if not cs:
            continue

        val = request.POST.get(key, '').strip()
        if not val:
            cs.allocated_teacher_id = None
            cs.teacher_id = None
        else:
            try:
                profile_id = int(val)
            except ValueError:
                continue
            profile = teacher_map.get(profile_id)
            if not profile:
                continue
            cs.allocated_teacher_id = profile.id
            cs.teacher_id = profile.user_id
        updated.append(cs)

    if updated:
        ClassSubject.objects.bulk_update(updated, ['allocated_teacher', 'teacher'], batch_size=100)

    messages.success(request, 'Тақсимоти дарсҳо ба омӯзгорон бо муваффақият сабт шуд.')
    return redirect('lesson_allocation')


def _school_sort_key(school):
    """Standard sort order used for the monitoring dashboard."""
    name_lower = (school.name or '').lower()
    type_lower = (school.type or '').lower()

    if 'идор' in type_lower:
        group_priority = 4
    elif 'томактаб' in type_lower:
        group_priority = 3
    elif 'лит' in type_lower or 'лиц' in type_lower:
        group_priority = 2
    elif 'гимн' in name_lower:
        group_priority = 1
        return (group_priority, 0, school.name)
    else:
        group_priority = 1

    nums = re.findall(r'\d+', school.name)
    num = int(nums[0]) if nums else 999999
    return (group_priority, num, school.name)


def _get_monitoring_stats():
    """Return the same statistics list used by the dashboard and the Excel export."""
    schools = sorted(School.objects.all(), key=_school_sort_key)
    stats = []
    for school in schools:
        school_num = get_school_number(school)
        zavuch_username = f'zavuch_{school_num}'
        zavuch_user = User.objects.filter(username=zavuch_username).first()
        logged_in = bool(zavuch_user and zavuch_user.last_login)
        stats.append({
            'school': school,
            'zavuch_username': zavuch_username,
            'zavuch_user': zavuch_user,
            'logged_in': logged_in,
            'class_count': Student.objects.filter(school=school).values('class_name').distinct().count(),
            'teacher_count': Teacher.objects.filter(school=school).count(),
            'student_count': Student.objects.filter(school=school).count(),
        })
    return stats


@login_required
def monitoring_dashboard(request):
    """Monitoring dashboard for superusers, staff, and school Zavuchs."""
    role = get_user_role(request.user)
    is_zavuch = (role and role.lower() == 'zavuch') or request.user.username.lower().startswith('zavuch_')
    if not (request.user.is_superuser or request.user.is_staff or is_zavuch):
        return redirect('dashboard')
    return render(request, 'portal/monitoring_dashboard.html', {'stats': _get_monitoring_stats()})


@login_required
def export_monitoring_excel(request):
    """Export the monitoring dashboard data as a styled Excel workbook."""
    if not request.user.is_superuser:
        return redirect('dashboard')

    stats = _get_monitoring_stats()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Назорати муассисаҳо'

    headers = [
        '№',
        'Муассиса',
        'Номи корбар',
        'Воридшавӣ',
        'Миқдори синфҳо',
        'Миқдори омӯзгорон',
        'Миқдори хонандагон',
    ]

    thin_side = Side(style='thin', color='000000')
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
    header_font = Font(bold=True, color='000000')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    for idx, row in enumerate(stats, start=2):
        ws.cell(row=idx, column=1, value=idx - 1).alignment = center_align
        ws.cell(row=idx, column=2, value=row['school'].name).alignment = left_align
        ws.cell(row=idx, column=3, value=row['zavuch_username']).alignment = left_align

        login_status = 'Фаъол' if row['logged_in'] else 'Ғайрифаъол'
        ws.cell(row=idx, column=4, value=login_status).alignment = center_align

        ws.cell(row=idx, column=5, value=row['class_count']).alignment = center_align
        ws.cell(row=idx, column=6, value=row['teacher_count']).alignment = center_align
        ws.cell(row=idx, column=7, value=row['student_count']).alignment = center_align

        for col in range(1, 8):
            ws.cell(row=idx, column=col).border = thin_border

    column_widths = [6, 45, 18, 16, 20, 22, 25]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    ws.row_dimensions[1].height = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    today = datetime.date.today().strftime('%Y-%m-%d')
    filename = f"Гузориши_Назорати_Муассисаҳо_{today}.xlsx"
    encoded_filename = urllib.parse.quote(filename)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f"attachment; filename*=utf-8''{encoded_filename}"
    return response


def google_verification(request):
    """Return the Google Search Console verification file content."""
    return HttpResponse(
        'google-site-verification: google3c6e6431fb434e83.html',
        content_type='text/html'
    )
